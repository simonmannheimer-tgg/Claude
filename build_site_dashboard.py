import json
import os
import sys
import csv
from datetime import date
from collections import defaultdict
from urllib.parse import urlsplit

_user_site = os.path.join(os.path.expanduser("~"), "AppData", "Roaming", "Python", "Python312", "site-packages")
if _user_site not in sys.path:
    sys.path.insert(0, _user_site)

# ── paths ────────────────────────────────────────────────────────────────────
HTML_PATH    = r"C:\Users\simonma\OneDrive - JB HI-FI Group Pty Ltd\Desktop\Content Cluster Diagram.html"
EXCEL_PATH   = r"C:\Users\simonma\OneDrive - JB HI-FI Group Pty Ltd\Desktop\internal_all iwth API data.xlsx"
INLINKS_PATH = r"C:\Users\simonma\OneDrive - JB HI-FI Group Pty Ltd\Desktop\inlinks.csv"
OUTLINKS_PATH= r"C:\Users\simonma\OneDrive - JB HI-FI Group Pty Ltd\Desktop\outlinks_filtered.csv"
DESKTOP      = r"C:\Users\simonma\OneDrive - JB HI-FI Group Pty Ltd\Desktop"
TODAY        = date.today().strftime("%Y%m%d")

OUT_JSON = os.path.join(DESKTOP, f"TGG_Site_Analysis_{TODAY}.json")
OUT_HTML = os.path.join(DESKTOP, f"TGG_Content_Dashboard_{TODAY}.html")

TGG_HOST = "thegoodguys.com.au"

def _num(v):
    if v is None or v == "": return None
    try:    return float(v)
    except: return None

def strip_params(url):
    p = urlsplit(str(url))
    return p._replace(query="", fragment="").geturl().rstrip("/")

def is_internal(url):
    return TGG_HOST in str(url)

# ── step 1: extract cluster/embedding data from SF diagram ───────────────────
print("Step 1/5: Extracting cluster assignments from Content Cluster Diagram...")
with open(HTML_PATH, "r", encoding="utf-8") as f:
    raw = f.read()

start = raw.find('[{"url":"https://')
if start < 0:
    print("ERROR: node data not found in HTML"); sys.exit(1)

nodes_raw, _ = json.JSONDecoder().raw_decode(raw, start)
del raw

cluster_map = {}   # url -> {cluster, x, y}
for n in nodes_raw:
    url = strip_params(n.get("url", ""))
    if url and is_internal(url):
        cluster_map[url] = {
            "cluster": int(n.get("cluster", 0)),
            "x":       round(n.get("x", 0), 4),
            "y":       round(n.get("y", 0), 4),
        }
del nodes_raw
print(f"  > {len(cluster_map):,} clustered URLs from diagram")

# ── step 2: count inlinks per destination URL (from inlinks.csv) ─────────────
print("Step 2/5: Counting inlinks per page from inlinks.csv...")
inlink_count  = defaultdict(int)   # url -> unique internal sources
inlink_sources = defaultdict(set)  # url -> set of source URLs (for unique count)

with open(INLINKS_PATH, "r", encoding="utf-8", newline="") as f:
    reader = csv.DictReader(f)
    for i, row in enumerate(reader, 1):
        frm = strip_params(row.get("From", ""))
        to  = strip_params(row.get("To",   ""))
        if is_internal(frm) and is_internal(to) and frm and to:
            inlink_sources[to].add(frm)
        if i % 1_000_000 == 0:
            print(f"  > {i:,} rows processed...")

for url, sources in inlink_sources.items():
    inlink_count[url] = len(sources)
del inlink_sources
print(f"  > Inlink counts computed for {len(inlink_count):,} URLs")

# ── step 3: count outlinks per source URL (from outlinks_filtered.csv) ───────
print("Step 3/5: Counting outlinks per page from outlinks_filtered.csv...")
outlink_count = defaultdict(int)   # url -> unique internal destinations

with open(OUTLINKS_PATH, "r", encoding="utf-8", newline="") as f:
    reader = csv.DictReader(f)
    for i, row in enumerate(reader, 1):
        frm = strip_params(row.get("From", ""))
        to  = strip_params(row.get("To",   ""))
        if is_internal(frm) and is_internal(to) and frm and to:
            outlink_count[frm] += 1
        if i % 1_000_000 == 0:
            print(f"  > {i:,} rows processed...")

print(f"  > Outlink counts computed for {len(outlink_count):,} URLs")

# ── step 4: read Excel for SF crawl + GSC metrics ───────────────────────────
print("Step 4/5: Reading Excel...")
import openpyxl
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

def sheet_to_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    return [dict(zip(header, row)) for row in rows[1:]]

gsc_data   = {}  # url -> GSC metrics
crawl_data = {}  # url -> crawl metrics

for sheet_name in wb.sheetnames:
    ws   = wb[sheet_name]
    rows = sheet_to_dicts(ws)
    if not rows: continue
    cols = set(rows[0].keys())

    url_col = next((c for c in cols if c.lower() in ("address","url")), None)
    if not url_col: continue

    has_gsc   = any(c.lower() == "clicks"      for c in cols)
    has_crawl = any("status" in c.lower()       for c in cols)

    clicks_col      = next((c for c in cols if c.lower() == "clicks"),      None)
    impressions_col = next((c for c in cols if c.lower() == "impressions"),  None)
    ctr_col         = next((c for c in cols if c.lower() == "ctr"),         None)
    position_col    = next((c for c in cols if c.lower() == "position"),     None)
    status_col      = next((c for c in cols if "status" in c.lower() and "code" in c.lower()), None) \
                      or next((c for c in cols if "status" in c.lower()),    None)
    title_col       = next((c for c in cols if "title"  in c.lower()),      None)
    h1_col          = next((c for c in cols if c.lower().startswith("h1")), None)
    words_col       = next((c for c in cols if "word"   in c.lower()),      None)
    ctype_col       = next((c for c in cols if "content type" in c.lower()),None)

    for row in rows:
        raw_url = str(row.get(url_col, "") or "").strip()
        if not raw_url: continue
        u = strip_params(raw_url)
        if has_gsc and u not in gsc_data:
            gsc_data[u] = {
                "clicks":      _num(row.get(clicks_col)),
                "impressions": _num(row.get(impressions_col)),
                "ctr":         _num(row.get(ctr_col)),
                "position":    _num(row.get(position_col)),
            }
        if has_crawl and u not in crawl_data:
            crawl_data[u] = {
                "status_code":  _num(row.get(status_col))                    if status_col else None,
                "title":        str(row.get(title_col,  "") or "").strip()[:120] if title_col else "",
                "h1":           str(row.get(h1_col,     "") or "").strip()[:120] if h1_col    else "",
                "word_count":   _num(row.get(words_col))                     if words_col  else None,
                "content_type": str(row.get(ctype_col,  "") or "").strip()   if ctype_col  else "",
            }

wb.close()
print(f"  > GSC: {len(gsc_data):,} URLs | Crawl: {len(crawl_data):,} URLs")

# ── step 5: join — all_urls = union of excel + diagram nodes ─────────────────
print("Step 5/5: Joining and building outputs...")

all_urls = set(crawl_data.keys()) | set(gsc_data.keys()) | set(cluster_map.keys())
print(f"  > {len(all_urls):,} unique URLs across all sources")

enriched = []
for url in all_urls:
    cl    = cluster_map.get(url, {})
    gsc   = gsc_data.get(url,   {})
    crawl = crawl_data.get(url, {})
    enriched.append({
        "url":          url,
        "cluster":      cl.get("cluster"),
        "x":            cl.get("x"),
        "y":            cl.get("y"),
        "inlink_count": inlink_count.get(url, 0),
        "outlink_count":outlink_count.get(url, 0),
        "clicks":       gsc.get("clicks"),
        "impressions":  gsc.get("impressions"),
        "ctr":          gsc.get("ctr"),
        "position":     gsc.get("position"),
        "status_code":  crawl.get("status_code"),
        "title":        crawl.get("title", ""),
        "h1":           crawl.get("h1", ""),
        "word_count":   crawl.get("word_count"),
        "content_type": crawl.get("content_type", ""),
    })

gsc_hits      = sum(1 for r in enriched if r["clicks"]      is not None)
crawl_hits    = sum(1 for r in enriched if r["status_code"] is not None)
cluster_hits  = sum(1 for r in enriched if r["cluster"]     is not None)
inlink_hits   = sum(1 for r in enriched if r["inlink_count"] > 0)
outlink_hits  = sum(1 for r in enriched if r["outlink_count"] > 0)
print(f"  > {len(enriched):,} pages | {cluster_hits:,} clustered | {gsc_hits:,} GSC | {crawl_hits:,} crawl | {inlink_hits:,} have inlinks | {outlink_hits:,} have outlinks")

# Save JSON
with open(OUT_JSON, "w", encoding="utf-8") as f:
    json.dump({
        "generated":      TODAY,
        "total_pages":    len(enriched),
        "clustered":      cluster_hits,
        "gsc_matched":    gsc_hits,
        "crawl_matched":  crawl_hits,
        "pages": enriched
    }, f, separators=(",", ":"))

size_mb = os.path.getsize(OUT_JSON) / 1024 / 1024
print(f"  > JSON saved: {OUT_JSON} ({size_mb:.2f} MB)")

# ── build cluster labels ─────────────────────────────────────────────────────
from collections import Counter, defaultdict as dd2

cluster_url_map = dd2(list)
for r in enriched:
    if r["cluster"] is not None:
        cluster_url_map[r["cluster"]].append(r["url"])

def infer_label(urls):
    segs = []
    for u in urls:
        path = u.replace("https://www.thegoodguys.com.au","").replace("https://thegoodguys.com.au","")
        parts = [p for p in path.split("/") if p]
        if parts:
            segs.append(parts[0].split("?")[0])
    if not segs: return "Other"
    top = Counter(segs).most_common(1)[0][0]
    return top.replace("-"," ").title()

cluster_labels = {cid: infer_label(urls) for cid, urls in cluster_url_map.items()}
cluster_labels[None] = "Unclustered"

for r in enriched:
    r["cluster_label"] = cluster_labels.get(r["cluster"], "Other")

# ── build HTML dashboard ──────────────────────────────────────────────────────
PALETTE = [
    "#4e79a7","#f28e2b","#e15759","#76b7b2","#59a14f",
    "#edc948","#b07aa1","#ff9da7","#9c755f","#bab0ac",
    "#1f77b4","#ff7f0e","#2ca02c","#d62728","#9467bd",
    "#8c564b","#e377c2","#7f7f7f","#bcbd22","#17becf",
    "#aec7e8","#ffbb78","#98df8a","#ff9896","#c5b0d5",
    "#c49c94","#f7b6d2","#c7c7c7","#dbdb8d","#9edae5",
]
unique_clusters = sorted({r["cluster"] for r in enriched if r["cluster"] is not None})
color_map = {str(cid): PALETTE[i % len(PALETTE)] for i, cid in enumerate(unique_clusters)}
color_map["null"] = "#444455"

# For pages without embedding coords, scatter them with a jittered grid so they still appear
import math, random
random.seed(42)
unclustered = [r for r in enriched if r["x"] is None]
if unclustered:
    # place them in a block to the right of the main cluster cloud
    all_x = [r["x"] for r in enriched if r["x"] is not None]
    max_x = max(all_x) if all_x else 0
    cols = max(1, int(math.sqrt(len(unclustered))))
    for i, r in enumerate(unclustered):
        r["x"] = round(max_x + 15 + (i % cols) * 2 + random.uniform(-0.5, 0.5), 4)
        r["y"] = round((i // cols) * 2 + random.uniform(-0.5, 0.5), 4)

def slug(url):
    return url.replace("https://www.thegoodguys.com.au","").replace("https://thegoodguys.com.au","") or "/"

js_nodes = []
for r in enriched:
    js_nodes.append({
        "url":  r["url"],
        "slug": slug(r["url"]),
        "c":    r["cluster"],
        "cl":   r["cluster_label"],
        "x":    r["x"],
        "y":    r["y"],
        "il":   r["inlink_count"],
        "ol":   r["outlink_count"],
        "clk":  r["clicks"],
        "imp":  r["impressions"],
        "pos":  r["position"],
        "sc":   r["status_code"],
        "title":r["title"][:80] if r["title"] else "",
        "wc":   r["word_count"],
        "ct":   r["content_type"],
    })

data_json     = json.dumps(js_nodes, separators=(",",":"))
colors_json   = json.dumps(color_map)
labels_json   = json.dumps({str(k): v for k, v in cluster_labels.items()})
legend_items  = [(str(cid), cluster_labels.get(cid,"Other"), len(cluster_url_map[cid])) for cid in unique_clusters]
legend_items.append(("null","Unclustered", len(unclustered)))

legend_html = ""
for cid, label, cnt in legend_items:
    col = color_map.get(cid, "#444455")
    legend_html += f'<div class="leg-item" data-cid="{cid}"><div class="leg-dot" style="background:{col}"></div><span>{label} ({cnt})</span></div>\n'

html_out = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>TGG Content Cluster Dashboard {TODAY}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0;font-family:system-ui,sans-serif}}
body{{background:#0f1117;color:#e0e0e0;height:100vh;display:flex;flex-direction:column;overflow:hidden}}
#top{{padding:8px 14px;background:#1a1d27;border-bottom:1px solid #2e3147;display:flex;flex-wrap:wrap;gap:8px;align-items:center}}
#top h1{{font-size:14px;color:#fff;white-space:nowrap;margin-right:4px}}
select,input[type=text]{{background:#252836;border:1px solid #3a3f5c;color:#e0e0e0;padding:4px 8px;border-radius:4px;font-size:12px}}
input[type=text]{{width:220px}}
#stats{{font-size:11px;color:#888;white-space:nowrap;margin-left:auto}}
#legend-wrap{{width:100%;display:flex;flex-wrap:wrap;gap:5px;padding-top:4px;border-top:1px solid #2e3147;margin-top:4px}}
.leg-item{{display:flex;align-items:center;gap:4px;font-size:11px;cursor:pointer;padding:2px 6px;border-radius:3px;border:1px solid transparent}}
.leg-item:hover{{border-color:#3a3f5c}}
.leg-item.muted{{opacity:.2}}
.leg-dot{{width:9px;height:9px;border-radius:50%;flex-shrink:0}}
#canvas-wrap{{flex:1;position:relative}}
canvas{{position:absolute;top:0;left:0;width:100%;height:100%}}
#tip{{position:absolute;pointer-events:none;background:rgba(15,17,28,.97);border:1px solid #3a3f5c;border-radius:6px;padding:10px 14px;font-size:12px;max-width:340px;display:none;z-index:99}}
.t-cluster{{font-size:13px;font-weight:700;margin-bottom:3px}}
.t-url{{color:#7eb8f7;word-break:break-all;font-size:11px;margin-bottom:5px}}
.t-title{{color:#ccc;font-size:11px;margin-bottom:5px;font-style:italic}}
.t-row{{display:flex;justify-content:space-between;gap:14px;padding:2px 0;border-bottom:1px solid #252836}}
.t-row:last-child{{border:none}}
.t-label{{color:#888}}.t-val{{color:#fff;font-weight:600}}
</style>
</head>
<body>
<div id="top">
  <h1>TGG Content Clusters</h1>
  <div><label style="font-size:11px;color:#aaa">Size&nbsp;</label>
  <select id="metric">
    <option value="il">Inlinks</option>
    <option value="clk">GSC Clicks</option>
    <option value="imp">Impressions</option>
    <option value="ol">Outlinks</option>
    <option value="wc">Word Count</option>
  </select></div>
  <input type="text" id="search" placeholder="Filter URL or title…">
  <span id="stats"></span>
  <div id="legend-wrap">{legend_html}</div>
</div>
<div id="canvas-wrap">
  <canvas id="cvs"></canvas>
  <div id="tip"></div>
</div>
<script>
const NODES={data_json};
const COLORS={colors_json};
const LABELS={labels_json};

const cvs=document.getElementById('cvs');
const ctx=cvs.getContext('2d');
const wrap=document.getElementById('canvas-wrap');
const tip=document.getElementById('tip');

let metric='il', filter='', muted=new Set();
let T={{k:1,tx:0,ty:0}}, drag=false, lastXY=null, hovered=null;

function resize(){{
  cvs.width=wrap.clientWidth; cvs.height=wrap.clientHeight; draw();
}}
window.addEventListener('resize',resize);

const xs=NODES.map(n=>n.x), ys=NODES.map(n=>n.y);
const extX=[Math.min(...xs),Math.max(...xs)];
const extY=[Math.min(...ys),Math.max(...ys)];

function proj(nx,ny){{
  const pad=50,W=cvs.width,H=cvs.height;
  const sx=(W-pad*2)/(extX[1]-extX[0]||1);
  const sy=(H-pad*2)/(extY[1]-extY[0]||1);
  const s=Math.min(sx,sy);
  const cx=(W-s*(extX[1]-extX[0]))/2;
  const cy=(H-s*(extY[1]-extY[0]))/2;
  return[cx+(nx-extX[0])*s, cy+(ny-extY[0])*s];
}}

function mval(n){{ const v=n[metric]; return(v==null)?0:+v; }}
function radii(nodes){{
  const vals=nodes.map(mval), mx=Math.max(...vals)||1;
  return nodes.map(n=>2+Math.sqrt(mval(n)/mx)*12);
}}

function visible(){{
  const lf=filter.toLowerCase();
  return NODES.filter(n=>
    !muted.has(String(n.c)) && !muted.has('null'===String(n.c)?'null':String(n.c)) &&
    !(n.c===null && muted.has('null')) &&
    (!lf||(n.slug||'').toLowerCase().includes(lf)||(n.title||'').toLowerCase().includes(lf))
  );
}}

function draw(){{
  const W=cvs.width,H=cvs.height;
  ctx.clearRect(0,0,W,H);
  const vn=visible(), rd=radii(vn);
  ctx.save();
  ctx.translate(T.tx,T.ty);
  ctx.scale(T.k,T.k);
  vn.forEach((n,i)=>{{
    const[px,py]=proj(n.x,n.y);
    const r=rd[i];
    const col=COLORS[String(n.c)??'null']||COLORS['null'];
    ctx.beginPath();
    ctx.arc(px,py,r,0,Math.PI*2);
    ctx.fillStyle=col+(n===hovered?'ff':'99');
    if(n===hovered){{ctx.shadowColor=col;ctx.shadowBlur=12;}}
    ctx.fill();
    ctx.shadowBlur=0;
  }});
  ctx.restore();
  document.getElementById('stats').textContent=
    vn.length.toLocaleString()+' of '+NODES.length.toLocaleString()+' pages';
}}

function fmt(v){{
  if(v==null||v==='')return'—';
  const n=+v;
  if(Number.isNaN(n))return String(v);
  if(n>=1000)return n.toLocaleString(undefined,{{maximumFractionDigits:0}});
  return n%1===0?String(n):n.toFixed(1);
}}

function showTip(n,ex,ey){{
  tip.style.display='block';
  const col=COLORS[String(n.c??'null')]||COLORS['null'];
  tip.innerHTML=`
    <div class="t-cluster" style="color:${{col}}">${{LABELS[String(n.c)]||LABELS['null']||'Cluster '+n.c}}</div>
    <div class="t-url">${{n.url}}</div>
    ${{n.title?`<div class="t-title">${{n.title}}</div>`:''}}
    <div class="t-row"><span class="t-label">Inlinks</span><span class="t-val">${{fmt(n.il)}}</span></div>
    <div class="t-row"><span class="t-label">Outlinks</span><span class="t-val">${{fmt(n.ol)}}</span></div>
    <div class="t-row"><span class="t-label">GSC Clicks</span><span class="t-val">${{fmt(n.clk)}}</span></div>
    <div class="t-row"><span class="t-label">Impressions</span><span class="t-val">${{fmt(n.imp)}}</span></div>
    <div class="t-row"><span class="t-label">Position</span><span class="t-val">${{fmt(n.pos)}}</span></div>
    ${{n.wc?`<div class="t-row"><span class="t-label">Words</span><span class="t-val">${{fmt(n.wc)}}</span></div>`:''}}
    ${{n.sc?`<div class="t-row"><span class="t-label">Status</span><span class="t-val">${{n.sc}}</span></div>`:''}}
    ${{n.ct?`<div class="t-row"><span class="t-label">Type</span><span class="t-val">${{n.ct}}</span></div>`:''}}
  `;
  const W=cvs.width,H=cvs.height,tw=340,th=tip.offsetHeight+20;
  tip.style.left=(ex+16+tw>W?ex-tw-8:ex+16)+'px';
  tip.style.top=(ey+th>H?H-th:ey)+'px';
}}

function hitTest(cx,cy){{
  const vn=visible(),rd=radii(vn);
  const mx=(cx-T.tx)/T.k, my=(cy-T.ty)/T.k;
  for(let i=vn.length-1;i>=0;i--){{
    const[px,py]=proj(vn[i].x,vn[i].y);
    if((mx-px)**2+(my-py)**2<=(Math.max(rd[i],6))**2) return vn[i];
  }}
  return null;
}}

cvs.addEventListener('mousemove',e=>{{
  const r=cvs.getBoundingClientRect(),cx=e.clientX-r.left,cy=e.clientY-r.top;
  if(drag&&lastXY){{T.tx+=cx-lastXY[0];T.ty+=cy-lastXY[1];lastXY=[cx,cy];draw();tip.style.display='none';return;}}
  const n=hitTest(cx,cy);
  hovered=n||null;
  cvs.style.cursor=n?'pointer':'grab';
  n?showTip(n,cx,cy):tip.style.display='none';
  draw();
}});
cvs.addEventListener('mousedown',e=>{{drag=true;const r=cvs.getBoundingClientRect();lastXY=[e.clientX-r.left,e.clientY-r.top];cvs.style.cursor='grabbing';}});
cvs.addEventListener('mouseup',()=>{{drag=false;lastXY=null;cvs.style.cursor='grab';}});
cvs.addEventListener('mouseleave',()=>{{drag=false;lastXY=null;tip.style.display='none';}});
cvs.addEventListener('click',e=>{{
  const r=cvs.getBoundingClientRect(),n=hitTest(e.clientX-r.left,e.clientY-r.top);
  if(n)window.open(n.url,'_blank');
}});
cvs.addEventListener('wheel',e=>{{
  e.preventDefault();
  const r=cvs.getBoundingClientRect(),cx=e.clientX-r.left,cy=e.clientY-r.top;
  const d=e.deltaY>0?0.85:1.18,nk=Math.min(Math.max(T.k*d,0.05),60);
  T.tx=cx-(cx-T.tx)*(nk/T.k); T.ty=cy-(cy-T.ty)*(nk/T.k); T.k=nk; draw();
}},{{passive:false}});

document.getElementById('metric').addEventListener('change',e=>{{metric=e.target.value;draw();}});
document.getElementById('search').addEventListener('input',e=>{{filter=e.target.value;draw();}});

document.querySelectorAll('.leg-item').forEach(el=>{{
  el.addEventListener('click',()=>{{
    const cid=el.dataset.cid;
    muted.has(cid)?muted.delete(cid):muted.add(cid);
    el.classList.toggle('muted');
    draw();
  }});
}});

resize();
</script>
</body>
</html>"""

with open(OUT_HTML, "w", encoding="utf-8") as f:
    f.write(html_out)

html_mb = os.path.getsize(OUT_HTML)/1024/1024
print(f"  > HTML saved: {OUT_HTML} ({html_mb:.2f} MB)")
print(f"\nDone.")
print(f"  Upload to Claude.ai: {OUT_JSON}")
print(f"  Open in browser:     {OUT_HTML}")
